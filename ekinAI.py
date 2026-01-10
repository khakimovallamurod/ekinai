import streamlit as st
import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split, cross_val_score, learning_curve
from sklearn.metrics import accuracy_score, classification_report
from imblearn.over_sampling import SMOTE
import plotly.express as px
import plotly.graph_objects as go
import io
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from PIL import Image as PILImage
import os

# Custom CSS for modern, attractive, thematic design: earth tones, responsive, centered, interactive elements
st.markdown("""
<style>
    /* Global styles */
    .stApp {
        background-color: white; /* Soft earth tone background */
        font-family: 'Roboto', sans-serif;
        max-width: 1200px; /* Make app window a bit smaller than full screen */
        margin: 0 auto; /* Center the app */
        padding: 1rem;
        border-radius: 12px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
    }
    h1, h2, h3 {
        text-align: center;
        color: #4a4a4a; /* Dark earth tone */
    }
    .stButton > button {
        background-color: #6b8e23; /* Olive green for theme */
        color: white;
        border-radius: 12px;
        padding: 1rem 4rem; /* Make button longer */
        font-size: 1.2rem;
        display: block;
        margin: 1rem auto;
        transition: all 0.3s ease;
    }
    .stButton > button:hover {
        background-color: #556b2f;
        transform: scale(1.05);
    }
    /* Download button similar to predict button */
    .stDownloadButton > button {
        background-color: #6b8e23;
        color: white;
        border-radius: 12px;
        padding: 1rem 4rem;
        font-size: 1.2rem;
        display: block;
        margin: 1rem auto;
        transition: all 0.3s ease;
    }
    .stDownloadButton > button:hover {
        background-color: #556b2f;
        transform: scale(1.05);
    }
    /* Input styles: attractive, modern */
    .stSelectbox, .stSlider {
        background-color: #fff;
        border: 1px solid #a9a9a9;
        border-radius: 8px;
        padding: 0.75rem;
        margin-bottom: 1.5rem;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        height: 60px !important; /* Uniform height */
        display: flex;
        align-items: center;
    }
    label {
        font-weight: bold;
        color: #397639; /* Thematic green */
        text-align: center;
        display: block;
    }
    /* Responsive columns */
    @media (max-width: 768px) {
        .stColumns {
            flex-direction: column;
        }
    }
    /* Logo centered */
    .logo-container {
        text-align: center;
    }
    /* Tables and dataframes */
    .stDataFrame {
        border: 1px solid #d2b48c; /* Tan border for theme */
        border-radius: 8px;
        overflow: hidden;
    }
    /* Modern scientific table styling */
    .stDataFrame table {
        font-family: 'Courier New', monospace; /* Scientific font */
        border-collapse: collapse;
        width: 100%;
    }
    .stDataFrame th, .stDataFrame td {
        border: 1px solid #ddd;
        padding: 8px;
        text-align: right;
    }
    .stDataFrame th {
        background-color: #f2f2f2;
        font-weight: bold;
    }
    /* Result container: interactive hover */
    .result-container {
        background-color: #e0d8c9; /* Light earth tone */
        border-radius: 12px;
        padding: 1.5rem;
        margin: 2rem 0;
        transition: box-shadow 0.3s ease;
    }
    .result-container:hover {
        box-shadow: 0 4px 16px rgba(0,0,0,0.15);
    }

    div .eng {
    text-align: center;
    
    }
    .best {
    font-size: 3rem;
        font-weight: bold;
        color: #42e3f5; 
        text-align: center;
        margin: 1rem 0;

    }
    .best-crop {
        font-size: 3rem;
        font-weight: bold;
        color: #228b22; /* Forest green */
        text-align: center;
        margin: 1rem 0;
        text-shadow: 1px 1px 2px rgba(0,0,0,0.1);
    }
    div .text {
    font-size: 3rem;
    text-align: center;
    margin-bottom: 2rem;
    }
    .ekin {
        font-weight: bold;
        
        color: #397639; /* Forest green */
        text-align: center;
    }
    .ai {
        font-weight: bold;
        
        color: #10eaf0; /* Blue for AI */
        text-align: center;
    }
    .matn {
        
        color: #000;
        text-align: center;
    }
    /* Expander for modals */
    .stExpander {
        background-color: #fff; /* Beige for sections */
        border-radius: 8px;
        margin-bottom: 1rem;
    }
    /* Custom slider styles: larger arrows, green theme, uniform size */
    .stSlider .css-1cpxqw2 {
        background-color: #6b8e23 !important; /* Green thumb (pointer) */
        width: 20px !important; /* Larger thumb */
        height: 20px !important;
        border-radius: 50% !important;
    }
    .stSlider .css-1cpxqw2:hover {
        background-color: #556b2f !important;
    }
    /* Arrows (steppers) larger and green */
    .stSlider button[data-testid="baseButton-secondary"] {
        background-color: #6b8e23 !important;
        color: white !important;
        font-size: 1.2rem !important;
        width: 40px !important;
        height: 40px !important;
        border-radius: 50% !important;
        margin: 0 5px !important;
    }
    .stSlider button[data-testid="baseButton-secondary"]:hover {
        background-color: #556b2f !important;
    }


  div[data-baseweb="select"] {
        cursor: pointer !important; /* sichqoncha olib borganda pointer */
        transition: 0.2s ease;
    }

    /* hover paytida biroz effekt berish (ixtiyoriy) */
    div[data-baseweb="select"]:hover {
        box-shadow: 0 0 4px #2b6c2b;
        border-radius: 6px;
    }

    

    /* Slider label styles */
    .slider-label {
        position: relative;
        display: inline-block;
        cursor: pointer;
        width: 100%;
        font-weight: bold;
        color: #397639; 
        text-align: left;
        margin-bottom: 5px;
    }

    /* Tooltip text */
    .tooltiptext {
        visibility: hidden;
        width: 220px;
        background-color: #555;
        color: #fff;
        text-align: center;
        border-radius: 6px;
        padding: 5px;
        position: absolute;
        z-index: 1;
        bottom: 125%; /* Position above */
        left: 50%;
        margin-left: -110px;
        opacity: 0;
        transition: opacity 0.3s;
        contain: content; /* Ensure tooltip stays within container */
    }
    /* Show tooltip on label hover */
    .slider-label:hover .tooltiptext {
        visibility: visible;
        opacity: 1;
    }
    /* Copyright section */
    .copyright {
        text-align: center;
        color: #005454;
        font-size: 1rem;
        margin-top: 2rem;
        border: 0.5px solid #eceeee;
        border-radius: 1rem;
        padding: 1.5rem;
        display: flex;
        justify-content: center;
        align-items: center;
    }

     .author {
        text-align: center;
        color: #005454;
        font-size: 1rem;
        display: flex;
        padding: 1rem;
        justify-content: center;
        align-items: center;
    }

     
    .copyright img {
        margin-right: 10px;
        width: 50px; /* Adjustable logo size */
    }

    .copyright .samdu {
        text-decoration: none;
        color: darkcyan;
        padding: 0.5rem;

    }

    .copyright .lab {
        
        text-decoration: none;
        color: #15803d;
        padding: 0.5rem;
         font-weight: 1.2rem;

    }

    .author .lab {
        text-decoration: none;
        color: #15803d;
        padding: 0.5rem;
        font-weight: bold;

    }

    .author .lab:hover {
        text-decoration: none;
        color: #24afff;
        padding: 0.5rem;
        font-weight: bold;

    }


    .copyright .lab:hover {
        text-decoration: none;
        color: #43d902;
        padding: 0.5rem;
        font-weight: 1.2rem;
        

    }

    .copyright .ekin2 {
        font-weight: bold;
        padding-left: 0.5rem;
        color: #397639; /* Forest green */
         
    }



    .author .ekin2 {

        padding-left: 0.5rem;
        color: #397639; /* Forest green */
        
    }
    .copyright .ai2 {
        font-weight: bold;
        color: #10eaf0; /* Blue for AI */
        
    }

    .author .ai2 {
      
        padding-left: 0.5rem;
        color: darkcyan 
        
    }

</style>
""", unsafe_allow_html=True)
# Logo (thematic placeholder, replace with actual if needed)
st.markdown('<div class="logo-container"><img src="https://raw.githubusercontent.com/abroraxatov1/ekinai/refs/heads/main/logos.jpg" alt="Logo" width="200"></div>', unsafe_allow_html=True)
# Dataset
DATA_URL = "ekin7.csv"
def parse_layer(value):
    if pd.isna(value):
        return np.nan
    value = str(value).strip().lower()
    if 'sm' in value or 'cm' in value:
        value = value.replace(' sm', '').replace(' cm', '')
    if '-' in value:
        parts = value.split('-')
        if len(parts) == 2:
            try:
                start = float(parts[0].strip())
                end = float(parts[1].strip())
                return (start + end) / 2
            except ValueError:
                pass
    try:
        return float(value)
    except ValueError:
        return np.nan
@st.cache_data
def load_and_preprocess():
    df = pd.read_csv(DATA_URL)
    df.columns = [col.strip() for col in df.columns]
    expected_columns = [
        'Qatlam (sm)', 'Mexanik tarkib', 'DNS (%)', 'Tuproq zichligi (g/cm³)',
        'pH', 'EC (mS/cm)', 'N (mg/kg)', 'P (mg/kg)', 'K (mg/kg)', 'Gumus (%)',
        'Mg (mg/kg)', 'S (mg/kg)', 'Zn (mg/kg)', 'Mn (mg/kg)', 'B (mg/kg)',
        'Fe (mg/kg)', 'Cu (mg/kg)', 'Mikroorganizmlar(CFU/g)', 'Ekin'
    ]
    for col in expected_columns:
        if col not in df.columns:
            df[col] = np.nan
    df['Ekin'] = df['Ekin'].fillna(df['Ekin'].mode()[0] if not df['Ekin'].mode().empty else 'Unknown').astype(str)
    if 'Qatlam (sm)' in df.columns:
        df['Qatlam (sm)'] = df['Qatlam (sm)'].apply(parse_layer)
        df['Qatlam (sm)'] = pd.to_numeric(df['Qatlam (sm)'], errors='coerce').fillna(df['Qatlam (sm)'].mean())
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    df[numeric_cols] = df[numeric_cols].fillna(df[numeric_cols].mean())
    categorical_cols = df.select_dtypes(exclude=[np.number]).columns
    for col in categorical_cols:
        if col != 'Ekin' and col != 'Namuna':
            df[col] = df[col].fillna(df[col].mode()[0] if not df[col].mode().empty else 'Unknown').astype(str)
    le_dict = {}
    for col in categorical_cols:
        if col != 'Ekin' and col != 'Namuna':
            le = LabelEncoder()
            df[col] = le.fit_transform(df[col])
            le_dict[col] = le
    le_crop = LabelEncoder()
    df['ekin_encoded'] = le_crop.fit_transform(df['Ekin'])
    crop_averages = df.groupby('Ekin')[numeric_cols].mean()
    return df, le_dict, le_crop, crop_averages
@st.cache_resource
def train_model(df):
    X = df.drop(['Namuna', 'Ekin', 'ekin_encoded'], axis=1, errors='ignore')
    y = df['ekin_encoded']
    for col in X.columns:
        X[col] = pd.to_numeric(X[col], errors='coerce').fillna(0)
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)
    smote = SMOTE(random_state=42, k_neighbors=5)
    X_train_res, y_train_res = smote.fit_resample(X_train, y_train)
    rf = RandomForestClassifier(
        n_estimators=100,
        max_depth=8,
        min_samples_split=5,
        min_samples_leaf=2,
        max_features='sqrt',
        random_state=42,
        class_weight='balanced'
    )
    rf.fit(X_train_res, y_train_res)
    train_acc = accuracy_score(y_train, rf.predict(X_train))
    test_acc = accuracy_score(y_test, rf.predict(X_test))
    cv_scores = cross_val_score(rf, X_train_res, y_train_res, cv=5, scoring='accuracy')
    cv_acc = cv_scores.mean()
    report = classification_report(y_test, rf.predict(X_test), target_names=le_crop.classes_, output_dict=True)
    feature_importances = pd.Series(rf.feature_importances_, index=X.columns).sort_values(ascending=False)
    # Compute learning curve
    train_sizes, train_scores, valid_scores = learning_curve(
        rf, X, y, train_sizes=np.linspace(0.1, 1.0, 5), cv=5, scoring='accuracy', n_jobs=-1, random_state=42
    )
    train_mean = np.mean(train_scores, axis=1)
    train_std = np.std(train_scores, axis=1)
    valid_mean = np.mean(valid_scores, axis=1)
    valid_std = np.std(valid_scores, axis=1)
    return rf, X.columns.tolist(), test_acc, cv_acc, train_acc, report, feature_importances, train_sizes, train_mean, train_std, valid_mean, valid_std
st.set_page_config(page_title="Ekin tavfsiyasi", layout="wide")
st.markdown(f'<div class="text"><span class="ekin"></span> <span class="ai"></span> <span class="matn"> <span class="ekin"> - Tuproq tarkibi asosida ekinlar hosildorligini aniqlashning </span> <span class="ai"> sun\'iy intellekt </span> <span class="ekin"> tizimi </span> </span></div>', unsafe_allow_html=True)
with st.spinner("Tizim ishga tushirilmoqda..."):
    df, le_dict, le_crop, crop_averages = load_and_preprocess()
    rf, feature_names, test_acc, cv_acc, train_acc, report, feature_importances, train_sizes, train_mean, train_std, valid_mean, valid_std = train_model(df)
# Center inputs in two columns instead of sidebar
st.subheader("Tuproq ma'lumotlarini kiriting")
# Tooltips for each property
tooltips = {
    "Qatlam (sm)": "Tuproq namunasi qatlam chuqurligi (sm)",
    "Mexanik tarkib": "Tuproqning mexanik tarkibi",
    "DNS (%)": "Tuproqning dala nam sig'imi ko'rsatkichi (%)",
    "Tuproq zichligi (g/cm³)": "Tuproqning zichlig (g/cm³)",
    "pH": "Tuproqning muhiti (kislotalik yoki ishqoriylik darajasi)",
    "EC (mS/cm)": "Tuproqning elektr o'tkazuvchanligi (mS/cm)",
    "N (mg/kg)": "Tuproqdagi umumiy azot miqdori (mg/kg)",
    "P (mg/kg)": "Tuproqdagi harakatchan fosfor miqdori (mg/kg)",
    "K (mg/kg)": "Tuproqdagi almashinuvchan kaliy miqdori (mg/kg)",
    "Gumus (%)": "Gumus - tuproqdagi organik moddalar ko'rsatgichi (%)",
    "Mg (mg/kg)": "Tuproqdagi magniy miqdori (mg/kg)",
    "S (mg/kg)": "Tuproqdagi oltingugurt miqdori (mg/kg)",
    "Zn (mg/kg)": "Tuproqdagi rux miqdori (mg/kg)",
    "Mn (mg/kg)": "Tuproqdagi marganes miqdori (mg/kg)",
    "B (mg/kg)": "Tuproqdagi bor miqdori (mg/kg)",
    "Fe (mg/kg)": "Tuproqdagi temir miqdori (mg/kg)",
    "Cu (mg/kg)": "Tuproqdagi mis miqdori (mg/kg)",
    "Mikroorganizmlar (CFU/g)": "Tuproqdagi mikroorganizmlar faolligi ko'rsatgichi (CFU/g)"
}
col1, col2 = st.columns(2)
with col1:
    st.markdown('<div class="slider-label" style="cursor: pointer; ">Qatlam (sm)<span class="tooltiptext">' + tooltips["Qatlam (sm)"] + '</span></div>', unsafe_allow_html=True)
    layer_options = ['0-20', '20-40', '40-60', '60-90']
    selected_layer = st.selectbox("", layer_options, label_visibility="collapsed")
    layer_value = (float(selected_layer.split('-')[0]) + float(selected_layer.split('-')[1])) / 2
    st.markdown('<div class="slider-label">Mexanik tarkib<span class="tooltiptext">' + tooltips["Mexanik tarkib"] + '</span></div>', unsafe_allow_html=True)
    mech_comp_encoded = 0
    if 'Mexanik tarkib' in le_dict:
        mech_classes = le_dict['Mexanik tarkib'].classes_
        mech_comp_str = st.selectbox("", mech_classes, label_visibility="collapsed")
        mech_comp_encoded = le_dict['Mexanik tarkib'].transform([mech_comp_str])[0]
    else:
        mech_comp_str = "Unknown"
    st.markdown('<div class="slider-label">DNS (%)<span class="tooltiptext">' + tooltips["DNS (%)"] + '</span></div>', unsafe_allow_html=True)
    dns = st.slider("", 1.0, 50.0, 25.0, label_visibility="collapsed", key="dns_slider")
    st.markdown('<div class="slider-label">Tuproq zichligi (g/cm³)<span class="tooltiptext">' + tooltips["Tuproq zichligi (g/cm³)"] + '</span></div>', unsafe_allow_html=True)
    density = st.slider("", 1.0, 3.0, 1.4, label_visibility="collapsed", key="density_slider")
    st.markdown('<div class="slider-label">pH<span class="tooltiptext">' + tooltips["pH"] + '</span></div>', unsafe_allow_html=True)
    ph = st.slider("", 4.0, 9.0, 6.5, label_visibility="collapsed", key="ph_slider")
    st.markdown('<div class="slider-label">EC (mS/cm)<span class="tooltiptext">' + tooltips["EC (mS/cm)"] + '</span></div>', unsafe_allow_html=True)
    ec = st.slider("", 0.1, 5.0, 1.0, label_visibility="collapsed", key="ec_slider")
    st.markdown('<div class="slider-label">N (mg/kg)<span class="tooltiptext">' + tooltips["N (mg/kg)"] + '</span></div>', unsafe_allow_html=True)
    nitrogen = st.slider("", 5.0, 200.0, 100.0, label_visibility="collapsed", key="nitrogen_slider")
    st.markdown('<div class="slider-label">P (mg/kg)<span class="tooltiptext">' + tooltips["P (mg/kg)"] + '</span></div>', unsafe_allow_html=True)
    phosphorus = st.slider("", 5.0, 100.0, 50.0, label_visibility="collapsed", key="phosphorus_slider")
    st.markdown('<div class="slider-label">K (mg/kg)<span class="tooltiptext">' + tooltips["K (mg/kg)"] + '</span></div>', unsafe_allow_html=True)
    potassium = st.slider("", 50.0, 500.0, 250.0, label_visibility="collapsed", key="potassium_slider")
with col2:
    st.markdown('<div class="slider-label">Gumus (%)<span class="tooltiptext">' + tooltips["Gumus (%)"] + '</span></div>', unsafe_allow_html=True)
    humus = st.slider("", 0.5, 5.0, 2.0, label_visibility="collapsed", key="humus_slider")
    st.markdown('<div class="slider-label">Mg (mg/kg)<span class="tooltiptext">' + tooltips["Mg (mg/kg)"] + '</span></div>', unsafe_allow_html=True)
    mg = st.slider("", 20.0, 500.0, 150.0, label_visibility="collapsed", key="mg_slider")
    st.markdown('<div class="slider-label">S (mg/kg)<span class="tooltiptext">' + tooltips["S (mg/kg)"] + '</span></div>', unsafe_allow_html=True)
    s = st.slider("", 5.0, 500.0, 50.0, label_visibility="collapsed", key="s_slider")
    st.markdown('<div class="slider-label">Zn (mg/kg)<span class="tooltiptext">' + tooltips["Zn (mg/kg)"] + '</span></div>', unsafe_allow_html=True)
    zn = st.slider("", 0.5, 10.0, 5.0, label_visibility="collapsed", key="zn_slider")
    st.markdown('<div class="slider-label">Mn (mg/kg)<span class="tooltiptext">' + tooltips["Mn (mg/kg)"] + '</span></div>', unsafe_allow_html=True)
    mn = st.slider("", 1.0, 60.0, 25.0, label_visibility="collapsed", key="mn_slider")
    st.markdown('<div class="slider-label">B (mg/kg)<span class="tooltiptext">' + tooltips["B (mg/kg)"] + '</span></div>', unsafe_allow_html=True)
    b = st.slider("", 0.1, 5.0, 2.0, label_visibility="collapsed", key="b_slider")
    st.markdown('<div class="slider-label">Fe (mg/kg)<span class="tooltiptext">' + tooltips["Fe (mg/kg)"] + '</span></div>', unsafe_allow_html=True)
    fe = st.slider("", 10.0, 400.0, 100.0, label_visibility="collapsed", key="fe_slider")
    st.markdown('<div class="slider-label">Cu (mg/kg)<span class="tooltiptext">' + tooltips["Cu (mg/kg)"] + '</span></div>', unsafe_allow_html=True)
    cu = st.slider("", 0.5, 10.0, 5.0, label_visibility="collapsed", key="cu_slider")
    st.markdown('<div class="slider-label">Mikroorganizmlar (CFU/g)<span class="tooltiptext">' + tooltips["Mikroorganizmlar (CFU/g)"] + '</span></div>', unsafe_allow_html=True)
    microorg = st.slider("", 1e5, 1e9, 1e8, label_visibility="collapsed", key="microorg_slider")
# Centered predict button
predict_button = st.button("Kiritish")
# Modern emojis for crops
crop_emojis = {
    "Bug'doy": "🌾",
    "Kartoshka": "🥔",
    "Loviya": "🫘",
    "Qalampir": "🌶️",
    "Makkajo'xori": "🌽",
    "Sabzi": "🥕",
    "Paxta": "☁️",
}
# Function to generate XLSX for single result with diagram image
def generate_xlsx(input_dict, df_probs, best_crop, fig):
    input_df = pd.DataFrame.from_dict(input_dict, orient='index', columns=['Qiymat'])
    input_df.index.name = 'Xususiyat'
    input_df.reset_index(inplace=True)
   
    probs_df = df_probs.copy()
   
    best_df = pd.DataFrame({'Eng mos ekin': [best_crop]})
   
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        input_df.to_excel(writer, sheet_name='Kiritilgan qiymatlar', index=False)
        probs_df.to_excel(writer, sheet_name='Ekin mosliklari', index=False)
        best_df.to_excel(writer, sheet_name='Eng mos ekin', index=False)
       
        # Add diagram as image to a new sheet if possible
        wb = writer.book
        img_sheet = wb.create_sheet('Moslik diagrammasi')
        try:
            img_data = io.BytesIO()
            fig.write_image(img_data, format='png')
            img_data.seek(0)
            img = PILImage.open(img_data)
            img_path = 'temp_diagram.png'
            img.save(img_path)
            img_openpyxl = Image(img_path)
            img_sheet.add_image(img_openpyxl, 'A1')
            os.remove(img_path)
        except Exception as e:
            # Skip adding image if error (e.g., kaleido not installed)
            pass
        
        bold = Font(bold=True)
        center = Alignment(horizontal='center')
        for sheet_name in wb.sheetnames:
            if sheet_name != 'Moslik diagrammasi':
                ws = wb[sheet_name]
                for cell in ws[1]:
                    cell.font = bold
                    cell.alignment = center
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column].width = adjusted_width
    output.seek(0)
    return output
# Interactive result appearance
if predict_button:
    input_dict = {
        'Qatlam (sm)': layer_value,
        'Mexanik tarkib': mech_comp_encoded,
        'DNS (%)': dns,
        'Tuproq zichligi (g/cm³)': density,
        'pH': ph,
        'EC (mS/cm)': ec,
        'N (mg/kg)': nitrogen,
        'P (mg/kg)': phosphorus,
        'K (mg/kg)': potassium,
        'Gumus (%)': humus,
        'Mg (mg/kg)': mg,
        'S (mg/kg)': s,
        'Zn (mg/kg)': zn,
        'Mn (mg/kg)': mn,
        'B (mg/kg)': b,
        'Fe (mg/kg)': fe,
        'Cu (mg/kg)': cu,
        'Mikroorganizmlar(CFU/g)': microorg
    }
    input_data = [input_dict.get(col, 0.0) for col in feature_names]
    input_df = pd.DataFrame([input_data], columns=feature_names)
    probs = rf.predict_proba(input_df)[0]
    crop_probs = {le_crop.classes_[i]: min(probs[i] * 100, 100) for i in range(len(le_crop.classes_))}
    df_probs = pd.DataFrame(list(crop_probs.items()), columns=['Ekin', 'Moslik (%)']).sort_values('Moslik (%)', ascending=False)
   
    best_crop = df_probs.iloc[0]['Ekin']
    best_prob = df_probs.iloc[0]['Moslik (%)']
    emoji = crop_emojis.get(best_crop, "🌾")
   
    # Display best crop with emojis below button
    st.markdown(f'<div class="eng"> <span class="best"> Eng mos ekin: </span> <span class="best-crop"> {best_crop} ({best_prob:.1f}%) {emoji}</span> </div>', unsafe_allow_html=True)
   
    with st.container():
        st.subheader("Ekinlarning moslik foizlari (0-100%)")
        st.table(df_probs)
        fig = px.bar(df_probs, x='Ekin', y='Moslik (%)', title="Ekin Mosligi (0-100% Diagrammasi)", color='Moslik (%)',
                     color_continuous_scale=["orange", "darkgreen", "green"], text='Moslik (%)')  # Changed to viridis for scientific look
        fig.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
        fig.update_layout(template="plotly_white", hovermode="x unified", font=dict(family="Arial", size=12, color="black"),
                          title_font=dict(size=20, color="darkgreen"), xaxis_title_font=dict(size=14), yaxis_title_font=dict(size=14))
        st.plotly_chart(fig, use_container_width=True)
       
        col1, col2 = st.columns(2, gap="medium")
        with col1:
            st.subheader(f"{best_crop} ekini rasmi")
            crop_images = {
                "Bug'doy": "https://i0.wp.com/razzanj.com/wp-content/uploads/2016/07/nature-landscape-nature-landscape-hd-image-download-wheat-farm-hd-wallpaper-notebook-background-wheat-farmers-wheat-farming-process-wheat-farming-in-kenya.jpg?ssl=1",
                "Kartoshka": "https://www.isaaa.org/kc/cropbiotechupdate/files/images/3172020111359PM.jpg",
                "Loviya": "https://cdn.britannica.com/24/122524-050-4593E7D1/Green-beans.jpg",
                "Qalampir": "https://img.freepik.com/premium-photo/red-chili-chilli-peppers-plant-garden-agricultural-plantation-farm-countryside-nonthaburi-thailand_258052-6029.jpg",
                "Makkajo'xori": "https://www.aces.edu/wp-content/uploads/2018/08/shutterstock_-Zeljko-Radojko_field-corn.jpg",
                "Sabzi": "https://ogden_images.s3.amazonaws.com/www.motherearthnews.com/images/2022/02/11110505/growing-carrots.jpg",
                "Paxta": "https://cdn.pixabay.com/photo/2014/02/13/12/56/cotton-crop-265312_1280.jpg",
            }
            if best_crop in crop_images:
                st.image(crop_images[best_crop], use_container_width=True, caption=f"{best_crop} ekini")
            else:
                st.info(f"{best_crop} uchun rasm topilmadi.")
        with col2:
            st.subheader(f"{best_crop}ga muxim xususiyatlar")
            top_features = feature_importances.head(6).index.tolist()
            input_values = pd.Series([input_dict.get(f, 0) for f in top_features], index=top_features)
            avg_values = crop_averages.loc[best_crop, top_features] if best_crop in crop_averages.index else pd.Series(0, index=top_features)
           
            # Grouped bar chart for better scientific comparison
            comparison_df = pd.DataFrame({
                'Xususiyat': top_features * 2,
                'Qiymat': list(input_values) + list(avg_values),
                'Turi': ['Kiritilgan qiymatlar'] * len(top_features) + [f"{best_crop} O'rtacha"] * len(top_features)
            })
            fig_bar = px.bar(comparison_df, x='Xususiyat', y='Qiymat', color='Turi', barmode='group',
                             color_discrete_sequence=['#1f77b4', '#2ca02c'])
            fig_bar.update_layout(hovermode="x unified", template="plotly_white")
            fig_bar.update_traces(hovertemplate="<b>%{x}</b><br>%{data.name}: %{y:.2f}")
            st.plotly_chart(fig_bar, use_container_width=True)
# Download button only for current result
if predict_button:
    xlsx = generate_xlsx(input_dict, df_probs, best_crop, fig)
    st.download_button(
        label="Natijani yuklash",
        data=xlsx,
        file_name="ekin_tavfsiyasi_natijasi.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# Bottom sections in expanders (modal-like)
with st.expander("Ma'lumotlar to'plami"):
    st.metric("Ma'lumotlar to'plami hajmi", f"{df.shape[0]} qator")
    st.write("**Ma'lumotlar to'plamining o'rtacha qiymatlari (min, max, mean, std, etc.):**")
    stats = df.describe()
    st.dataframe(stats.style.format("{:.2f}"), use_container_width=True) # Removed highlight_max
    st.write("**Siniflar taqqoslovi (balanslanmagan holda):**")
    class_counts = df['Ekin'].value_counts()
    fig_class = px.bar(class_counts, x=class_counts.index, y=class_counts.values, title="Siniflar",
                       color=class_counts.values, color_continuous_scale="sunset")  # Changed to blues for distinction
    fig_class.update_layout(template="plotly_white", font=dict(family="Arial", size=12, color="black"),
                            title_font=dict(size=20, color="darkblue"), xaxis_title_font=dict(size=14), yaxis_title_font=dict(size=14))
    st.plotly_chart(fig_class, use_container_width=True)
with st.expander("Model ma'lumotlari"):
    col_metrics1, col_metrics2, col_metrics3 = st.columns(3)
    with col_metrics1:
        st.metric("Train aniqligi", f"{train_acc:.2%}")
    with col_metrics2:
        st.metric("Test aniqligi", f"{test_acc:.2%}")
    with col_metrics3:
        st.metric("Kross-validatsiya aniqligi", f"{cv_acc:.2%}")
   
    # Learning curve plot - made clearer, no fills, thicker lines
    fig_lc = go.Figure()
    fig_lc.add_trace(go.Scatter(x=train_sizes, y=train_mean, mode='lines+markers', name='Training accuracy', line=dict(color='blue', width=3)))
    fig_lc.add_trace(go.Scatter(x=train_sizes, y=valid_mean, mode='lines+markers', name='Valid accuracy', line=dict(color='red', width=3)))
    fig_lc.update_layout(title="O'rganish chizig'i", xaxis_title='Training examples', yaxis_title='Score', template='plotly_white', hovermode='x unified',
                         font=dict(family="Arial", size=12, color="black"), title_font=dict(size=20, color="black"),
                         xaxis_title_font=dict(size=14), yaxis_title_font=dict(size=14))
    st.plotly_chart(fig_lc, use_container_width=True)
with st.expander("Kiritilgan qiymatlar"):
    input_summary = {
        "Qatlam (sm)": layer_value,
        "Mexanik tarkib": mech_comp_str,
        "DNS (%)": dns,
        "Tuproq zichligi (g/cm³)": density,
        "pH": ph,
        "EC (mS/cm)": ec,
        "N (mg/kg)": nitrogen,
        "P (mg/kg)": phosphorus,
        "K (mg/kg)": potassium,
        "Gumus (%)": humus,
        "Mg (mg/kg)": mg,
        "S (mg/kg)": s,
        "Zn (mg/kg)": zn,
        "Mn (mg/kg)": mn,
        "B (mg/kg)": b,
        "Fe (mg/kg)": fe,
        "Cu (mg/kg)": cu,
        "Mikroorganizmlar (CFU/g)": microorg
    }
    summary_list = list(input_summary.items())
    half = len(summary_list) // 2
    col1, col2 = st.columns(2)
    with col1:
        st.table(pd.DataFrame(summary_list[:half], columns=["Xususiyat", "Qiymat"]))
    with col2:
        st.table(pd.DataFrame(summary_list[half:], columns=["Xususiyat", "Qiymat"]))
# Copyright at the bottom with optional logo
logo_url = "https://www.samdu.uz/new/images/SamDU%20logo%20full%201.png"  # Replace with your logo URL or leave empty
if logo_url:
    st.markdown(f'<div class="copyright"><img src="https://www.samdu.uz/new/images/SamDU%20logo%20full%201.png" alt="Logo"> <a href="https://www.samdu.uz/uz" class="samdu"> Samarqand davlat universiteti    </a> <a href="http://ai-lab.sampc.uz/" class="lab"> "Sun\'iy intellekt labaratoriyasi" </a> <span class="ishlab"> tomonidan ishlab chiqilgan. © 2025 </span> <span class="ekin2">Ekin</span> <span class="ai2"> AI</span>.</div>', unsafe_allow_html=True)
else:

    st.markdown(f'<div class="copyright"> <a href="https://www.samdu.uz/uz" class="samdu"> Samarqand davlat universiteti  </a> <a href="http://ai-lab.sampc.uz/" class="lab"> "Sun\'iy intellekt labaratoriyasi" </a> <span class="ishlab"> tomonidan ishlab chiqilgan. © 2025 </span> <span class="ekin2">Ekin</span> <span class="ai2"> AI</span>.</div>', unsafe_allow_html=True)

st.markdown(f'<div class="author">  Muallif: <a href="https://t.me/axatov_a" class="lab"> Axatov Abror </a> <span class="ekin2">Tel: </span> <span class="ai2"> +998(99)-590-52-65 </span></div>', unsafe_allow_html=True)





































